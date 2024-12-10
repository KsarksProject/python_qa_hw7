import csv
import os
import zipfile

import PyPDF2
import pytest
from openpyxl import load_workbook


@pytest.fixture(scope="module")
def zip_file_with_files():
    files_to_archive = [
        'tmp/Black_Hat_Python.pdf',
        'tmp/sample2.xlsx',
        'tmp/sample4.csv'
    ]
    zip_file_name = 'tmp/test_files.zip'

    with zipfile.ZipFile(zip_file_name, 'w') as archive:
        for file in files_to_archive:
            archive.write(file, os.path.basename(file))
    yield zip_file_name

    os.remove(zip_file_name)


def read_pdf(file):
    reader = PyPDF2.PdfReader(file)
    return ''.join(page.extract_text() for page in reader.pages)


def read_excel(file):
    from io import BytesIO

    file_bytes = BytesIO(file.read())
    workbook = load_workbook(file_bytes)
    sheet = workbook.active
    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    if len(data) > 1:
        keys = data[0]
        return [dict(zip(keys, row)) for row in data[1:]]
    return data


def read_csv(file):
    csv_content = []
    file = file.read().decode('utf-8').splitlines()
    reader = csv.DictReader(file)
    for row in reader:
        csv_content.append(row)
    return csv_content


def test_pdf_content(zip_file_with_files):
    with zipfile.ZipFile(zip_file_with_files, 'r') as archive:
        with archive.open('Black_Hat_Python.pdf') as file:
            content = read_pdf(file)
    assert "Black Hat Python" in content, "Содержимое PDF файла некорректно"


def test_excel_content(zip_file_with_files):
    with zipfile.ZipFile(zip_file_with_files, 'r') as archive:
        with archive.open('sample2.xlsx') as file:
            content = read_excel(file)
    assert any(row.get("Creator") == "John Doe" for row in content), "Содержимое Excel файла некорректно"


def test_csv_content(zip_file_with_files):
    with zipfile.ZipFile(zip_file_with_files, 'r') as archive:
        with archive.open('sample4.csv') as file:
            content = read_csv(file)
    assert any(row.get("Game Number") == "1" for row in content), "Содержимое CSV файла некорректно"
